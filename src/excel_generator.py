# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class ContentCalendarExcelGenerator:
    """Generate professional Excel content calendars"""

    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Define colors
        self.colors = {
            'header': 'B4D3E8',  # Light blue
            'pillar1': 'FFE5B4',  # Peach
            'pillar2': 'D4F1D4',  # Light green
            'pillar3': 'E8D4F1',  # Light purple
            'pillar4': 'FFD4D4',  # Light red
        }

    def create_monthly_calendar_tab(self, content_pieces, brand_name, month):
        """Create the main monthly calendar view"""
        ws = self.wb.create_sheet("Monthly Calendar")

        # Title
        ws['A1'] = f"{brand_name} - Content Calendar"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = month
        ws['A2'].font = Font(size=12, color='666666')

        # Headers
        headers = ['Date', 'Day', 'Content Title', 'Channel', 'Format', 'Pillar', 'Status', 'Notes']
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        header_font = Font(bold=True, color='000000')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Content rows
        for idx, piece in enumerate(content_pieces, 5):
            ws.cell(row=idx, column=1, value=piece.get('suggested_date', ''))
            ws.cell(row=idx, column=2, value=self._get_day_of_week(piece.get('suggested_date', '')))
            ws.cell(row=idx, column=3, value=piece.get('title', ''))
            ws.cell(row=idx, column=4, value=piece.get('channel', ''))
            ws.cell(row=idx, column=5, value=piece.get('format', ''))
            ws.cell(row=idx, column=6, value=piece.get('pillar', ''))
            ws.cell(row=idx, column=7, value='Draft')  # Default status
            ws.cell(row=idx, column=8, value='')  # Empty notes

            # Color code by pillar
            pillar_num = self._extract_pillar_number(piece.get('pillar', ''))
            if pillar_num:
                pillar_color = self.colors.get(f'pillar{pillar_num}', 'FFFFFF')
                for col in range(1, 9):
                    ws.cell(row=idx, column=col).fill = PatternFill(
                        start_color=pillar_color,
                        end_color=pillar_color,
                        fill_type='solid'
                    )

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30

        return ws

    def create_content_details_tab(self, content_pieces):
        """Create detailed content specifications tab"""
        ws = self.wb.create_sheet("Content Details")

        # Headers for Content Details
        detail_headers = [
            "ID", "Week", "Title", "Pillar", "Channel", "Format",
            "Key Message", "Description", "CTA", "Effort",
            "Engagement", "SEO Keyword", "Notes"
        ]

        ws.append(detail_headers)

        # Format headers
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add content details data
        for piece in content_pieces:
            detail_row = [
                piece.get('content_id', ''),
                piece.get('week', ''),
                piece.get('title', ''),
                piece.get('pillar', ''),
                piece.get('channel', ''),
                piece.get('format', ''),
                piece.get('key_message', ''),
                piece.get('description', ''),
                piece.get('call_to_action', ''),
                piece.get('effort_level', ''),
                piece.get('engagement_potential', ''),
                piece.get('seo_keyword', ''),
                piece.get('execution_notes', '')
            ]
            ws.append(detail_row)

        # Set column widths for Content Details
        column_widths_details = [5, 6, 30, 15, 12, 15, 25, 35, 20, 10, 12, 20, 30]

        for idx, width in enumerate(column_widths_details, start=1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[column_letter].width = width

        # Wrap text for description, key message, and notes columns
        for row_idx in range(2, len(content_pieces) + 2):
            for col_idx in [7, 8, 9, 13]:  # Key Message, Description, CTA, Notes columns
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        return ws

    def create_weekly_checklist_tab(self, content_pieces):
        """Create weekly checklist view"""
        ws = self.wb.create_sheet("Weekly Checklist")

        checklist_headers = ["Week", "Content ID", "Title", "Status", "Notes"]
        ws.append(checklist_headers)

        # Format headers
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Group by week
        weeks = {}
        for piece in content_pieces:
            week = piece.get('week', 1)
            if week not in weeks:
                weeks[week] = []
            weeks[week].append(piece)

        # Add checklist data
        current_row = 2
        for week in sorted(weeks.keys()):
            for piece in weeks[week]:
                checklist_row = [
                    f"Week {week}",
                    piece.get('content_id', ''),
                    piece.get('title', ''),
                    "Pending",
                    ""
                ]
                ws.append(checklist_row)
                current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

        return ws

    def create_metrics_tab(self, content_pieces, success_metrics=None):
        """Create metrics tracking tab"""
        ws = self.wb.create_sheet("Metrics Tracker")

        # Title
        ws['A1'] = "Content Performance Metrics"
        ws['A1'].font = Font(size=14, bold=True)

        # Success Metrics section
        ws['A3'] = "Target Metrics"
        ws['A3'].font = Font(size=12, bold=True)

        if success_metrics:
            row = 4
            for metric in success_metrics:
                ws.cell(row=row, column=1, value='•')
                ws.cell(row=row, column=2, value=metric)
                row += 1

        # Performance tracking table
        row = row + 2
        ws.cell(row=row, column=1, value="Content Performance Tracking")
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1

        headers = ['Content ID', 'Title', 'Views', 'Engagement', 'Clicks', 'Conversions', 'Notes']
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        row += 1

        # Empty rows for tracking
        for piece in content_pieces:
            ws.cell(row=row, column=1, value=piece.get('content_id', ''))
            ws.cell(row=row, column=2, value=piece.get('title', ''))
            # Leave other columns empty for manual tracking
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30

        return ws

    def _get_day_of_week(self, date_string):
        """Extract day of week from date string"""
        # Simple extraction - assumes format like "January 15, 2025 (Monday)"
        if '(' in date_string and ')' in date_string:
            return date_string[date_string.find('(')+1:date_string.find(')')]
        return ''

    def _extract_pillar_number(self, pillar_text):
        """Extract pillar number for color coding"""
        if 'Pillar 1' in pillar_text or 'pillar 1' in pillar_text.lower():
            return 1
        elif 'Pillar 2' in pillar_text or 'pillar 2' in pillar_text.lower():
            return 2
        elif 'Pillar 3' in pillar_text or 'pillar 3' in pillar_text.lower():
            return 3
        elif 'Pillar 4' in pillar_text or 'pillar 4' in pillar_text.lower():
            return 4
        return None

    def save(self, filepath):
        """Save the workbook"""
        self.wb.save(filepath)
        return filepath


def generate_content_calendar_xlsx(brand_name, month, content_pieces, success_metrics=None, output_path='content_calendar.xlsx'):
    """
    Generate a comprehensive Excel content calendar

    Args:
        brand_name: Name of the brand
        month: Month (e.g., "January 2025")
        content_pieces: List of content piece dictionaries
        success_metrics: Optional list of success metrics
        output_path: Where to save the file

    Returns:
        Path to saved file
    """
    gen = ContentCalendarExcelGenerator()

    gen.create_monthly_calendar_tab(content_pieces, brand_name, month)
    gen.create_content_details_tab(content_pieces)
    gen.create_weekly_checklist_tab(content_pieces)
    gen.create_metrics_tab(content_pieces, success_metrics)

    return gen.save(output_path)
